from openpyxl import load_workbook
from datetime import datetime, timedelta, date as dtdate
from collections import defaultdict
from typing import List, Dict, Tuple, TypedDict
import openai
from dotenv import load_dotenv

# Load environment variables (for OpenAI API key)
load_dotenv()


# Define a TypedDict for entries
class Entry(TypedDict):
    date: dtdate
    hours: float
    description: str


def get_week_start(date: dtdate) -> dtdate:
    """Returns the Monday of the ISO week for a given date."""
    return date - timedelta(days=date.weekday())


def summarize_with_llm(descriptions: List[str]) -> str:
    """Use LLM to generate a concise summary of tasks."""
    try:
        prompt = (
            "Summarize these work tasks into one brief professional sentence for an invoice. Keep mentions of the specifics. "
            "Avoid technical jargon and repetition:\n- " + "\n- ".join(descriptions)
        )

        response = openai.chat.completions.create(
            model="gpt-4-turbo",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2,
        )
        if response.choices[0].message.content:
            return response.choices[0].message.content.strip()
        else:
            raise Exception("no response")
    except Exception as e:
        print(f"LLM Error: {str(e)}")
        return "; ".join(descriptions)  # Fallback to original method


def summarize_hours_by_week(sheet_name: str, filename: str) -> List[Entry]:
    """Summarize hours by week from an Excel sheet."""
    try:
        wb = load_workbook(filename=filename, data_only=True)
    except FileNotFoundError:
        print("File does not exist.")
        exit(1)

    try:
        sheet = wb[sheet_name]
    except KeyError:
        print(f"Sheet '{sheet_name}' not found.")
        exit(1)

    entries: List[Entry] = []
    current_date: dtdate | None = None

    for row in sheet.iter_rows(min_row=5):
        date_cell = row[0]
        hours_cell = row[1]
        desc_cell = row[2]

        if date_cell.value and isinstance(date_cell.value, datetime):
            current_date = date_cell.value.date()
        else:
            if (
                hours_cell.value is not None
                and desc_cell.value is not None
                and current_date is not None
            ):
                entries.append(
                    {
                        "date": current_date,
                        "hours": float(hours_cell.value),
                        "description": desc_cell.value.strip(),
                    }
                )

    weekly_entries: Dict[Tuple[int, int], List[Entry]] = defaultdict(list)
    for entry in entries:
        year, week, _ = entry["date"].isocalendar()
        weekly_entries[(year, week)].append(entry)

    summaries: List[Entry] = []
    for key in sorted(weekly_entries.keys()):
        week_entries = weekly_entries[key]
        total_hours = sum(entry["hours"] for entry in week_entries)
        week_start = get_week_start(week_entries[0]["date"])

        # Collect unique descriptions
        seen = set()
        unique_descriptions: List[str] = []
        for entry in week_entries:
            desc = entry["description"]
            if desc not in seen:
                seen.add(desc)
                unique_descriptions.append(desc)

        # Generate LLM summary
        llm_summary = summarize_with_llm(unique_descriptions)

        summaries.append(
            Entry(date=week_start, hours=total_hours, description=llm_summary)
        )

    return summaries
